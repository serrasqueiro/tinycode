#-*- coding: utf-8 -*-
# xcelent.py  (c)2020  Henrique Moreira

"""
Basic Excel content (openpyxl wrapper)
"""

# pylint: disable=no-self-use, missing-function-docstring

import openpyxl


def main_test():
    for num in (1, 10, 26, 27, 28, 3423):
        # https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
        letters = num_to_column_letter(num)
        backnum = column_index_from_letter(letters)
        print(f"column index {num}={backnum} has letter(s): {letters}")


class Libre():
    """ Libre abstract class """
    name = ""

    def get_name(self):
        return self.name


class Xcel(Libre):
    """ Xcel and LibreOffice workbooks wrapper class """
    workbook = None
    _sheets = None

    """ Xcel workbook """
    def __init__(self, wb, name=""):
        self.name = name
        self.workbook = wb
        self._init_sheets()

    def _init_sheets(self):
        self._sheets = dict_from_sheets(self.workbook)

    def get_sheet(self, idx_name):
        """ get_sheet() from column index, or column letter name """
        if isinstance(idx_name, int):
            idx = idx_name
            assert idx > 0
            sheet = self._sheets.get(f"@{idx}")
        else:
            sheet = self._sheets.get(idx_name)
        # Returns openpyxl.worksheet.worksheet.Worksheet
        return sheet


class Xsheet(Libre):
    """ Xcel sheet """
    _sheet = None
    rows = []

    def __init__(self, sheet, name=""):
        self.name = name
        self._sheet = sheet
        self.rows = self._from_sheet(sheet)

    def _from_sheet(self, sheet):
        # pylint: disable=unnecessary-comprehension
        rows = [row for row in sheet]
        if rows:
            self.column_refs = self._guess_columns(rows[0])
        else:
            self.column_refs = dict()
        return rows

    def _guess_columns(self, first) -> dict:
        res = dict()
        #cells = [text for text in first]
        idx, letter = 1, "A"
        letters = "@"
        for cell in first:
            letters += letter
            astr = cell.value if isinstance(cell.value, str) else letter
            res[letter] = astr
            res[idx] = astr
            letter = increment_column_letter(letter)
            idx += 1
        res["@letters"] = letters
        return res


def dict_from_sheets(wbk) -> dict:
    """ Returns a dictionary with all sheets """
    dct = {"@sheetnames": wbk.sheetnames,
           }
    idx = 0
    for name in wbk.sheetnames:
        idx += 1
        assert name
        assert not name.startswith("@")
        dct[name] = wbk[name]
        dct[f"@{idx}"] = wbk[name]
    return dct


def euro_value(num) -> str:
    assert isinstance(num, float)
    return "{:.2f}".format(float(num))


def increment_column_letter(letter) -> str:
    # The simplest form would be:
    #	assert letter <= "Z"
    #	res = chr(ord(letter) + 1)
    num = column_index_from_letter(letter)
    return num_to_column_letter(num+1)


def num_to_column_letter(num) -> str:
    assert isinstance(num, int)
    assert num >= 1
    assert num < 10000	# just on the sanity side...
    letters = openpyxl.utils.cell.get_column_letter(num)
    return letters

def column_index_from_letter(letters) -> int:
    assert isinstance(letters, str)
    num = openpyxl.utils.cell.column_index_from_string(letters)
    return num


# Main script
if __name__ == "__main__":
    print("Import filing.xcelent !")
    main_test()
